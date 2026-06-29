"""Write metadata spreadsheets for a folder tree (recursive).

Produces two CSVs (open directly in Excel, no third-party deps):
  <out>_folders.csv  one row per folder, size = total of everything beneath it
  <out>_files.csv    one row per file
Each top-level subtree is walked in its own thread.
"""
import csv
import os
import sys
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path


FOLDER_COLS = ["folder", "direct_files", "total_files", "direct_size_bytes",
               "total_size_bytes", "total_size_mb", "file_types", "created", "modified", "error"]
FILE_COLS = ["path", "folder", "name", "type", "size_bytes", "size_mb", "created", "modified", "error"]


def _dt(ts: float) -> str:
    return datetime.fromtimestamp(ts).isoformat(sep=" ", timespec="seconds")


def human_size(n: float) -> str:
    n = float(n)
    for unit in ("B", "KB", "MB", "GB", "TB"):
        if n < 1024:
            return f"{n:.0f} {unit}" if unit == "B" else f"{n:.1f} {unit}"
        n /= 1024
    return f"{n:.1f} PB"


def _created(st: os.stat_result) -> str:
    # st_birthtime = true creation time (Win/macOS, py3.12+); st_ctime fallback elsewhere
    return _dt(getattr(st, "st_birthtime", st.st_ctime))


def scan_tree(root: str) -> tuple[list[dict], list[dict]]:
    """Walk one subtree; return (folder_rows, file_rows). Recursive sizes roll up to parents."""
    folders, files = [], []
    rec_size = defaultdict(int)
    rec_count = defaultdict(int)
    rec_types = defaultdict(lambda: defaultdict(int))

    for dirpath, _dirnames, filenames in os.walk(root, topdown=False):  # bottom-up so children roll into parents
        direct_size = direct_count = 0
        for fn in filenames:
            fp = os.path.join(dirpath, fn)
            try:
                st = os.stat(fp)
            except OSError as e:
                files.append({"path": fp, "error": str(e)})
                continue
            ext = os.path.splitext(fn)[1].lower() or "(no ext)"
            direct_size += st.st_size
            direct_count += 1
            rec_types[dirpath][ext] += 1
            files.append({
                "path": fp, "folder": dirpath, "name": fn, "type": ext,
                "size_bytes": st.st_size, "size_mb": round(st.st_size / 1_048_576, 3),
                "created": _created(st), "modified": _dt(st.st_mtime),
            })

        rec_size[dirpath] += direct_size
        rec_count[dirpath] += direct_count

        try:
            st = os.stat(dirpath)
            folders.append({
                "folder": dirpath,
                "direct_files": direct_count,
                "total_files": rec_count[dirpath],
                "direct_size_bytes": direct_size,
                "total_size_bytes": rec_size[dirpath],
                "total_size_mb": round(rec_size[dirpath] / 1_048_576, 2),
                "file_types": ", ".join(f"{e}:{n}" for e, n in sorted(rec_types[dirpath].items())),
                "created": _created(st),
                "modified": _dt(st.st_mtime),
            })
        except OSError as e:
            folders.append({"folder": dirpath, "error": str(e)})

        parent = os.path.dirname(dirpath)  # roll this dir's recursive totals up
        rec_size[parent] += rec_size[dirpath]
        rec_count[parent] += rec_count[dirpath]
        for e, n in rec_types[dirpath].items():
            rec_types[parent][e] += n

    return folders, files


def scan_parallel(base: str, max_workers: int = 32, progress=None) -> tuple[list[dict], list[dict]]:
    """Scan a whole tree, fanning out across top-level subfolders (hides network latency).

    progress(done, total) is called as each top-level subfolder finishes. Includes base's
    own direct files and a roll-up row for base itself.
    """
    base = str(base)
    folders, files = [], []
    try:
        entries = list(os.scandir(base))
    except OSError as e:
        return [{"folder": base, "error": str(e)}], []

    subdirs = [e.path for e in entries if e.is_dir(follow_symlinks=False)]
    direct_size = direct_count = 0
    for e in entries:
        if not e.is_file(follow_symlinks=False):
            continue
        try:
            st = e.stat(follow_symlinks=False)
        except OSError as ex:
            files.append({"path": e.path, "error": str(ex)})
            continue
        direct_size += st.st_size
        direct_count += 1
        files.append({
            "path": e.path, "folder": base, "name": e.name,
            "type": os.path.splitext(e.name)[1].lower() or "(no ext)",
            "size_bytes": st.st_size, "size_mb": round(st.st_size / 1_048_576, 3),
            "created": _created(st), "modified": _dt(st.st_mtime),
        })

    with ThreadPoolExecutor(max_workers=max_workers) as pool:
        futures = {pool.submit(scan_tree, sd): sd for sd in subdirs}
        for i, fut in enumerate(as_completed(futures), 1):
            try:
                f_rows, file_rows = fut.result()
                folders += f_rows
                files += file_rows
            except Exception as e:
                folders.append({"folder": futures[fut], "error": str(e)})
            if progress:
                progress(i, len(subdirs))

    # base's own roll-up row, computed from everything collected under it
    own_files = [f for f in files if "size_bytes" in f]
    types = defaultdict(int)
    for f in own_files:
        types[f["type"]] += 1
    total_size = sum(f["size_bytes"] for f in own_files)
    try:
        st = os.stat(base)
        folders.append({
            "folder": base,
            "direct_files": direct_count,
            "total_files": len(own_files),
            "direct_size_bytes": direct_size,
            "total_size_bytes": total_size,
            "total_size_mb": round(total_size / 1_048_576, 2),
            "file_types": ", ".join(f"{e}:{n}" for e, n in sorted(types.items())),
            "created": _created(st),
            "modified": _dt(st.st_mtime),
        })
    except OSError as e:
        folders.append({"folder": base, "error": str(e)})
    return folders, files


def _children_roots(by_path: dict) -> tuple[defaultdict, list]:
    """From {folderpath: row}, derive parent->subfolders and the top-level (root) paths.
    ponytail: adding both a folder and its descendant as scan roots would list it twice; don't."""
    children: defaultdict[str, list] = defaultdict(list)
    roots = []
    for p in by_path:
        parent = os.path.dirname(p.rstrip("\\/"))
        if parent in by_path and parent != p:
            children[parent].append(p)
        else:
            roots.append(p)
    return children, roots


def _tree_order(folders: list[dict], files: list[dict]):
    """Rebuild the folder hierarchy from flat scan rows. Yields (kind, depth, record, fullpath)
    depth-first: each folder, then its direct files, then its subfolders."""
    folder_by_path = {f["folder"]: f for f in folders if "error" not in f}
    files_by_dir: defaultdict[str, list] = defaultdict(list)
    for fi in files:
        if "error" not in fi:
            files_by_dir[fi["folder"]].append(fi)
    children, roots = _children_roots(folder_by_path)
    roots.sort(key=str.lower)
    for kids in children.values():
        kids.sort(key=str.lower)

    def emit(fp, depth):
        yield ("folder", depth, folder_by_path[fp], fp)
        for fi in sorted(files_by_dir.get(fp, []), key=lambda x: x.get("name", "").lower()):
            yield ("file", depth + 1, fi, fi["path"])
        for c in children.get(fp, []):
            yield from emit(c, depth + 1)

    for r in roots:
        yield from emit(r, 0)


XLSX_MAX_ROWS = 1_048_576  # Excel hard limit per worksheet (incl. header)


def write_xlsx(path: str, folders: list[dict], files: list[dict]):
    """Readable workbook:
      'Tree'  - cascading folders (Excel outline groups), readable sizes. The 'Bytes (sum-safe)'
                column has values only on FILE rows, so summing it gives the true total once
                (folder rows are rollups for reading, never summed -> no double counting).
      'Files' - flat list of every file, fully summable for pivots.
    Sheets that exceed Excel's row limit overflow into '<name> (2)', '(3)', ...
    """
    # ponytail: openpyxl normal mode holds all cells in memory; fine into low millions of rows.
    # If a share is huge enough to OOM, switch the Files sheet to write_only mode.
    from openpyxl import Workbook
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE  # control chars Excel/XML reject
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    def clean(v):
        return ILLEGAL_CHARACTERS_RE.sub("", v) if isinstance(v, str) else v

    wb = Workbook()
    bold = Font(bold=True)
    folder_fill = PatternFill("solid", fgColor="E8EEF7")
    used_default = [False]

    def add_sheet(title, cols, outline=False):
        ws = wb.active if not used_default[0] else wb.create_sheet()
        used_default[0] = True
        ws.title = title
        ws.append(cols)
        for c in range(1, len(cols) + 1):
            ws.cell(1, c).font = bold
        ws.freeze_panes = "A2"
        if outline:
            ws.sheet_properties.outlinePr.summaryBelow = False  # +/- control on the folder row
        return ws

    # ---------- Tree sheet(s): cascading ----------
    tcols = ["Name", "Type", "Size", "Bytes (sum-safe)", "Items", "Created", "Modified", "Full path"]
    twidths = [len(c) for c in tcols]
    tsheets, part = [], 1
    ws = add_sheet("Tree", tcols, outline=True)
    tsheets.append(ws)
    for kind, depth, rec, fp in _tree_order(folders, files):
        if ws.max_row >= XLSX_MAX_ROWS:  # roll over before exceeding the limit
            part += 1
            ws = add_sheet(f"Tree ({part})", tcols, outline=True)
            tsheets.append(ws)
        is_folder = kind == "folder"
        name = fp if depth == 0 else (os.path.basename(fp.rstrip("\\/")) or fp)
        if is_folder:
            vals = [clean(name), "Folder", human_size(rec.get("total_size_bytes", 0)),
                    None, rec.get("total_files", ""), rec.get("created", ""),
                    rec.get("modified", ""), clean(fp)]
        else:
            vals = [clean(name), rec.get("type", ""), human_size(rec.get("size_bytes", 0)),
                    rec.get("size_bytes", 0), "", rec.get("created", ""),
                    rec.get("modified", ""), clean(fp)]
        ws.append(vals)
        ri = ws.max_row
        ws.cell(ri, 1).alignment = Alignment(indent=min(depth, 14))  # cascading indent
        ws.row_dimensions[ri].outline_level = min(depth, 7)          # Excel max outline depth = 7
        if is_folder:
            for c in range(1, len(tcols) + 1):
                ws.cell(ri, c).font = bold
                ws.cell(ri, c).fill = folder_fill
        for i, v in enumerate(vals):
            extra = depth * 2 if i == 0 else 0
            twidths[i] = max(twidths[i], len(str("" if v is None else v)) + extra)

    # ---------- Files sheet(s): flat ----------
    fwidths = [len(c) for c in FILE_COLS]
    fsheets, part = [], 1
    ws = add_sheet("Files", FILE_COLS)
    fsheets.append(ws)
    for fi in files:
        if ws.max_row >= XLSX_MAX_ROWS:
            part += 1
            ws = add_sheet(f"Files ({part})", FILE_COLS)
            fsheets.append(ws)
        vals = [clean(fi.get(c, "")) for c in FILE_COLS]
        ws.append(vals)
        for i, v in enumerate(vals):
            fwidths[i] = max(fwidths[i], len(str(v)))

    for sheets, widths in ((tsheets, twidths), (fsheets, fwidths)):
        for s in sheets:
            for i, w in enumerate(widths, 1):
                s.column_dimensions[get_column_letter(i)].width = min(w + 2, 70)

    wb.save(path)


def write_growth_xlsx(path: str, folders: list[dict], files: list[dict]):
    """One sheet: every folder at every depth, cascading, with size, % of total storage,
    file count, and created/modified dates. Folders only (no file rows)."""
    from openpyxl import Workbook
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
    from openpyxl.chart import BarChart, Reference
    from openpyxl.formatting.rule import DataBarRule
    from openpyxl.styles import Alignment, Font, PatternFill

    def clean(v):
        return ILLEGAL_CHARACTERS_RE.sub("", v) if isinstance(v, str) else v

    vfolders = [f for f in folders if "error" not in f]
    vfiles = [f for f in files if "error" not in f]
    total_size = sum(f.get("size_bytes", 0) for f in vfiles)
    pct = (lambda b: (b / total_size) if total_size else 0)

    by_path = {f["folder"]: f for f in vfolders}
    children, roots = _children_roots(by_path)
    wb = Workbook()
    h2 = Font(bold=True, size=12)
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="305496")
    label_font = Font(bold=True)
    bar = lambda ws, col, r0, r1: ws.conditional_formatting.add(
        f"{col}{r0}:{col}{r1}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color="5B9BD5"))

    def header(ws, row, cols):
        for i, c in enumerate(cols, 1):
            cell = ws.cell(row, i, c)
            cell.font = hdr_font
            cell.fill = hdr_fill
        return row + 1

    # ---- Folder Breakdown: every folder at every depth (the only sheet) ----
    brows: list[tuple] = []

    def walk(p, dep):
        fc = by_path[p]
        name = clean(p) if dep == 0 else clean(os.path.basename(p.rstrip("\\/")) or p)
        brows.append((dep, name, fc.get("total_size_bytes", 0), fc.get("total_files", 0),
                      clean(fc.get("file_types", "")), fc.get("created", ""), fc.get("modified", "")))
        for c in sorted(children.get(p, []), key=lambda c: -by_path[c].get("total_size_bytes", 0)):
            walk(c, dep + 1)

    for r in sorted(roots, key=lambda x: -by_path[x].get("total_size_bytes", 0)):
        walk(r, 0)

    cols = ["Folder", "Size", "% of total storage", "Files", "File types", "Created", "Modified"]
    widths = (("A", 60), ("B", 14), ("C", 16), ("D", 10), ("E", 36), ("F", 20), ("G", 20))
    sheets = []

    def new_sheet(n):
        sh = wb.active if n == 1 else wb.create_sheet()
        sh.title = "Folder Breakdown" if n == 1 else f"Folder Breakdown ({n})"
        sh["A1"] = "Folder breakdown - every folder, all depths"
        sh["A1"].font = h2
        sh.sheet_properties.outlinePr.summaryBelow = False
        header(sh, 3, cols)
        sh.freeze_panes = "A4"
        for col, w in widths:  # fixed widths: autosize would be O(rows) on a huge folder list
            sh.column_dimensions[col].width = w
        sh.sheet_view.showGridLines = False
        sheets.append(sh)
        return sh

    part = 1
    ws = new_sheet(1)
    for dep, name, sz, fcount, ftypes, created, modified in brows:
        if ws.max_row >= XLSX_MAX_ROWS:  # huge archives roll into Folder Breakdown (2), (3), ...
            part += 1
            ws = new_sheet(part)
        ws.append([name, human_size(sz), pct(sz), fcount, ftypes, created, modified])
        ri = ws.max_row
        ws.cell(ri, 1).alignment = Alignment(indent=min(dep * 2, 14))  # cascading indent
        ws.cell(ri, 3).number_format = "0.0%"
        ws.cell(ri, 4).number_format = "#,##0"
        ws.row_dimensions[ri].outline_level = min(dep, 7)              # Excel max outline depth = 7
        if dep == 0:
            for cc in range(1, len(cols) + 1):
                ws.cell(ri, cc).font = label_font
    for sh in sheets:
        if sh.max_row >= 4:
            bar(sh, "C", 4, sh.max_row)

    # ---- Template: you type patients & procedures; Excel formulas + chart do the rest ----
    tpl = wb.create_sheet("Template")
    tpl.sheet_view.showGridLines = False
    input_fill = PatternFill("solid", fgColor="FFF2CC")  # yellow = type your number here
    tpl["A1"] = "Cohort Summary"
    tpl["A1"].font = Font(bold=True, size=16)
    tpl["A2"] = "Type into the yellow cells (B5, B6). Everything else updates automatically."
    tpl["A2"].font = Font(italic=True, size=9, color="808080")
    rows = [  # (label, value-or-formula, number_format, is_input)
        ("Total imaging volume (GB)", round(total_size / 1_000_000_000, 2), "#,##0.00", False),
        ("Number of patients", None, "#,##0", True),
        ("Number of procedures", None, "#,##0", True),
        ("Procedures per patient", '=IF(B5,B6/B5,"")', "0.0", False),
        ("Avg data per patient (GB)", '=IF(B5,B4/B5,"")', "#,##0.00", False),
        ("Avg data per procedure (GB)", '=IF(B6,B4/B6,"")', "#,##0.00", False),
    ]
    for i, (label, val, fmt, is_input) in enumerate(rows, start=4):
        tpl.cell(i, 1, label).font = Font(bold=True)
        c = tpl.cell(i, 2)
        if val is not None:
            c.value = val
        c.number_format = fmt
        if is_input:
            c.fill = input_fill
    tpl.column_dimensions["A"].width = 30
    tpl.column_dimensions["B"].width = 16

    chart = BarChart()
    chart.type = "col"
    chart.title = "Average data (GB)"
    chart.legend = None
    chart.height, chart.width = 7, 12
    chart.add_data(Reference(tpl, min_col=2, min_row=8, max_row=9))   # avg per patient / per procedure
    chart.set_categories(Reference(tpl, min_col=1, min_row=8, max_row=9))
    tpl.add_chart(chart, "D4")

    wb.move_sheet(tpl, -wb.worksheets.index(tpl))  # put Template first
    wb.active = 0
    wb.save(path)


def _write(path: str, cols: list[str], rows: list[dict]):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:  # utf-8-sig so Excel reads unicode right
        w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def main(base: str = ".", out: str = "folder_metadata"):
    base_path = Path(base)
    if not base_path.exists():
        sys.exit(f"Path not found: {base_path}\n"
                 f"For a network host use a UNC share, e.g. \\\\10.10.100.130\\ShareName "
                 f"(run `net view \\\\10.10.100.130` to list shares).")
    all_folders, all_files = scan_parallel(str(base_path),
                                            progress=lambda d, t: print(f"  {d}/{t} top folders", end="\r"))

    all_folders.sort(key=lambda r: r["folder"].lower())
    all_files.sort(key=lambda r: r["path"].lower())

    _write(f"{out}_folders.csv", FOLDER_COLS, all_folders)
    _write(f"{out}_files.csv", FILE_COLS, all_files)
    print(f"Wrote {len(all_folders)} folders -> {out}_folders.csv")
    print(f"Wrote {len(all_files)} files   -> {out}_files.csv")


if __name__ == "__main__":
    main(*sys.argv[1:])  # usage: python folder_meta.py [base_dir] [out_prefix]


def test_scan():
    import tempfile
    with tempfile.TemporaryDirectory() as tmp:
        top = Path(tmp) / "top"
        nested = top / "nested"
        nested.mkdir(parents=True)
        (top / "a.txt").write_text("hello")        # 5 bytes
        (nested / "b.py").write_text("x=1")        # 3 bytes
        folders, files = scan_tree(str(top))
        assert len(files) == 2
        top_row = next(r for r in folders if r["folder"] == str(top))
        assert top_row["total_size_bytes"] == 8     # recursive: includes nested
        assert top_row["direct_size_bytes"] == 5    # direct: excludes nested
        assert top_row["total_files"] == 2
        assert ".py:1" in top_row["file_types"]
        print("test_scan ok")
